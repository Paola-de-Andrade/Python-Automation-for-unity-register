import pandas as pd
import tkinter as tk
from tkinter import simpledialog
from tkinter import messagebox

ROOT = tk.Tk()

ROOT.withdraw()
# the input dialog
versao_base = simpledialog.askstring(title="Versão atual da planilha de cadastro",
                                  prompt="Qual versão da planilha você quer usar? Ex.: v3")
#versao_base_nova = simpledialog.askstring(title="Versão nova da planilha de cadastro",
                                  #prompt="Qual versão da planilha você quer salvar ao terminar? Ex.: v4")

path = "Cadastro UC SEFIC_"+str(versao_base)+".xlsx"
#path = r"C:\Users\2003305\Documents\Cadastro UC SEFIC_"+str(versao_base)+".xlsx"

planilha = pd.read_excel(path)
planilha = planilha.T.set_index(0).T
planilha = planilha.iloc[: , :23]
# display(planilha)

#pip install openpyxl==2.6.3
#pip install selenium
#pip install time
#pip install datetime
#pip install tkinter
#pip install pandas
#pip install pywin32
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from datetime import date
import openpyxl
from datetime import datetime
from openpyxl.styles import colors
from openpyxl.styles import Font, Color 
from tkinter import messagebox
import win32com.client
  
link_login = "http://sefic.cpfl.com.br/SEFICad"
link_clientes = "http://sefic.cpfl.com.br/SEFIC/ModuloGestao/Cliente"
link_cadastroUC = "http://sefic.cpfl.com.br/SEFIC/ModuloGestao/Cliente/Inserir"
today = date.today()
# dd/mm/YYYY
d1 = today.strftime("%d/%m/%Y")
colunaCodigo = planilha['Seu Código']
linha=0
k=0
outlook = win32com.client.Dispatch('outlook.application')

#iteração para buscar fazer login no SEFIC mesmo se cachê estiver cheio e ficar dando erro
while k <100:
    try:
        driver = webdriver.Chrome()
        driver.get(link_login)
        time.sleep(5)
        teste_home = driver.find_element_by_xpath("/html/body/div/nav/div/div/div[3]/div[2]/ul/li[1]/a")
        break
    except:
        driver.close()
        time.sleep(20)
        k=k+1
        continue

k=0
        
#para cada célula de UC que está presente        
for cell in colunaCodigo:
    #pegar apenas as células não vazias
    if cell == cell:
        confirm = planilha.iloc[linha]['Confirmação']
        
        #pegar apenas as células da coluna de Confirmação que estão vazias
        if confirm != confirm:
            driver.get(link_clientes)
            UC = planilha.iloc[linha]['Seu Código']
            
            #pesquisar se aquela UC já existe
            driver.find_element_by_xpath("/html/body/div/div/div[1]/div[1]/div[2]/form/div[1]/div[3]/div/input").send_keys(Keys.CONTROL + 'a', Keys.BACKSPACE)
            driver.find_element_by_xpath("/html/body/div/div/div[1]/div[1]/div[2]/form/div[1]/div[3]/div/input").send_keys (UC)
            botao = driver.find_element_by_xpath('/html/body/div/div/div[1]/div[1]/div[2]/form/div[2]/button')
            botao.click()
            
            #o XPath abaixo corresponde à primeira coluna da primeira linha da tabela que aparece após a pesquisa
            #se não encontrar resultados, corresponderá a "Nenhum item foi encontrado."
            #se encontrar resultados, corresponderá a distribuidora do primeiro resultado encontrado
            text = driver.find_element_by_xpath("/html/body/div/div/div[1]/div[2]/div[2]/table/tbody/tr/td").text
            #se houver "item" no text, significa que não encontrou resultados, e retornará a posição de "item" em text
            #se não houver "item" no text, significa que encontrou resultados, e retornará -1
            validation = text.find("item")
            
            i=0
            #se encontrou resultados, iteraremos até o final da tabela até acharmos o texto "X registros encontrados"
            #e assim saberemos quantos registros foram encontrados para colocar isso no Excel
            while validation == -1:
                i=i+1
                text = driver.find_element_by_xpath("/html/body/div/div/div[1]/div[2]/div[2]/table/tbody/tr["+str(i)+"]/td").text
                validation = text.find("Registro")
            
            #se houver "item" no text, significa que não encontrou resultados, e retornará True
            #se não houver "item" no text, significa que encontrou resultados, e retornará False
            nao_cadastrado = "item" in text  
            
            #caso precisem haver zeros a serem completados na UC
            if len(str(UC))<10:
                zeros_faltantes = 10-len(str(UC))
                UC = str(UC)+zeros_faltantes*"0"
                
            #se não encontrou resultados, entrará no cadastro de UC
            if nao_cadastrado == True:
                driver.get(link_cadastroUC)
                time.sleep(5)
                #print("entrou pra fazer o cadastro da linha ="+str(linha))
                Nome = planilha.iloc[linha]['Nome/Razão Social']
                CNPJ = planilha.iloc[linha]['CPF/CNPJ']
                Logradouro = planilha.iloc[linha]['Logradouro']
                Numero_endereco = planilha.iloc[linha]['Número']
                CEP = planilha.iloc[linha]['CEP']
                Bairro= planilha.iloc[linha]['Bairro']
                Complemento = planilha.iloc[linha]['Complemento']
                if confirm != confirm:
                    Complemento = ""
                Email = planilha.iloc[linha]['E-mail']
                Codigo_Distribuidora = int(planilha.iloc[linha]['Código Distribuidora'])
                Codigo_Estado = int(planilha.iloc[linha]['Código Estado'])
                Codigo_Cidade = int(planilha.iloc[linha]['Código Cidade'])
                Codigo_Tensao = int(planilha.iloc[linha]['Código Tensão'])
                Codigo_Classe = int(planilha.iloc[linha]['Código Classe'])
                Codigo_Modalidade = int(planilha.iloc[linha]['Código Modalidade'])
                file_path = str(planilha.iloc[linha]['Caminho da fatura na rede'])
                
                #acima pegou os valores das variáveis e abaixo fará o preenchimento na página de cadastro
                driver.find_element_by_xpath("/html/body/div[1]/div/div/div/div[2]/form/div[1]/div[2]/div/input").send_keys (UC)
                driver.find_element_by_xpath("/html/body/div[1]/div/div/div/div[2]/form/div[1]/div[5]/div/input").send_keys (str(Nome))
                driver.find_element_by_xpath("/html/body/div[1]/div/div/div/div[2]/form/div[1]/div[6]/div/input").send_keys (str(Nome))
                driver.find_element_by_xpath("/html/body/div[1]/div/div/div/div[2]/form/div[1]/div[3]/div/input").send_keys (str(CNPJ))
                driver.find_element_by_xpath("/html/body/div[1]/div/div/div/div[2]/form/div[1]/div[11]/div/input").send_keys (str(Email))
                driver.find_element_by_xpath("/html/body/div[1]/div/div/div/div[2]/form/div[2]/div[5]/div/input").send_keys (str(Logradouro))
                driver.find_element_by_xpath("/html/body/div[1]/div/div/div/div[2]/form/div[2]/div[6]/div/input").send_keys (str(Numero_endereco))
                driver.find_element_by_xpath("/html/body/div[1]/div/div/div/div[2]/form/div[2]/div[9]/div/input").send_keys (str(CEP))
                driver.find_element_by_xpath("/html/body/div[1]/div/div/div/div[2]/form/div[2]/div[8]/div/input").send_keys (str(Bairro))
                driver.find_element_by_xpath("/html/body/div[1]/div/div/div/div[2]/form/div[2]/div[7]/div/input").send_keys (str(Complemento))
                
                ElementTensao = Select(driver.find_element_by_xpath("/html/body/div[1]/div/div/div/div[2]/form/div[1]/div[9]/div/select"))
                ElementTensao.select_by_value(str(Codigo_Tensao))

                ElementClasse = Select(driver.find_element_by_xpath("/html/body/div[1]/div/div/div/div[2]/form/div[1]/div[7]/div/select"))
                ElementClasse.select_by_value(str(Codigo_Classe))

                ElementModalidade = Select(driver.find_element_by_xpath("/html/body/div[1]/div/div/div/div[2]/form/div[1]/div[8]/div/select"))
                ElementModalidade.select_by_value(str(Codigo_Modalidade))
                
                ElementDistribuidora = Select(driver.find_element_by_xpath("/html/body/div[1]/div/div/div/div[2]/form/div[2]/div[2]/div/select"))
                ElementDistribuidora.select_by_value(str(Codigo_Distribuidora))

                ElementEstado = Select(driver.find_element_by_xpath("/html/body/div[1]/div/div/div/div[2]/form/div[2]/div[3]/div/select"))
                ElementEstado.select_by_value(str(Codigo_Estado))
                
                #tempo a ser esperado para que carregue as cidades após selecionar o Estado
                time.sleep(10)
                
                ElementCidade = Select(driver.find_element_by_xpath("/html/body/div[1]/div/div/div/div[2]/form/div[2]/div[4]/div/select"))
                ElementCidade.select_by_value(str(Codigo_Cidade))

                fileinput = driver.find_element_by_xpath("/html/body/div[1]/div/div/div/div[2]/form/div[2]/div[11]/div/input")
                fileinput.send_keys(file_path)
                
                #tempo colocado para não haja erro na inserção de data
                time.sleep(2)
                
                driver.find_element_by_xpath("/html/body/div[1]/div/div/div/div[2]/form/div[2]/div[1]/div/input").click()
                driver.find_element_by_xpath("/html/body/div[1]/div/div/div/div[2]/form/div[2]/div[1]/div/input").send_keys (d1)
                
                #tempo colocado para não haja erro na inserção de data
                time.sleep(2)
                
                botao = driver.find_element_by_xpath('/html/body/div[1]/div/div/div/div[2]/form/div[2]/div[12]/div/button')
                botao.click()

                #tempo colocado para carregar a página após submeter cadastro de UC
                time.sleep(10)
                
                #fará nova pesquisa procurando se aquela UC, já com os zeros completos, foi cadastrada
                driver.find_element_by_xpath("/html/body/div/div/div[2]/div[1]/div[2]/form/div[1]/div[3]/div/input").send_keys(Keys.CONTROL + 'a', Keys.BACKSPACE)
                driver.find_element_by_xpath("/html/body/div/div/div[2]/div[1]/div[2]/form/div[1]/div[3]/div/input").send_keys (UC)
                botao = driver.find_element_by_xpath('/html/body/div/div/div[2]/div[1]/div[2]/form/div[2]/button')
                botao.click()

                text = driver.find_element_by_xpath("/html/body/div/div/div[1]/div[2]/div[2]/table/tbody/tr/td").text
                validation = text.find("item")
                
                i=0
                #se encontrou resultados, iteraremos até o final da tabela até acharmos o texto "X registros encontrados"
                #e assim saberemos quantos registros foram encontrados para colocar isso no Excel                
                while validation == -1:
                    i=i+1
                    text = driver.find_element_by_xpath("/html/body/div/div/div[1]/div[2]/div[2]/table/tbody/tr["+str(i)+"]/td").text
                    validation = text.find("Registro")

                ja_cadastrado_checagem = "1 Registro" in text  
                
                #atualizar planilha com a confirmação de que foi cadastrado
                if ja_cadastrado_checagem == True:  
                    book = openpyxl.load_workbook(path)
                    sheet = book.active
                    fontStyle = Font(name="Calibri", size=12, color=colors.BLACK)
                    celula_confirmacao = sheet.cell(row=linha+3, column=23)
                    now = datetime.now()
                    data_cadastro = now.strftime("Cadastrado em "+"%d/%m/%Y "+"às"+" %H:%M:%S")
                    celula_confirmacao.value = data_cadastro
                    celula_confirmacao.font = fontStyle
                    book.save(path)
                    k=k+1
                    book.close()
                    
                    #enviar e-mail de confirmação de cadastro após preenchimento na planilha
                    mail = outlook.CreateItem(0)
                    mail.SendUsingAccount = 'peecpfl@cpfl.com.br'
                    mail.To = Email
                    mail.Subject = 'Cadastro da Unidade Consumidora '+ str(UC)+' – Chamada Pública de Projetos CPFL 2021'
                    mail.HTMLBody = '<p><font face="Calibri" ><p>Prezado(a) cliente '+ str(Nome)+',</p> <p>Seu cadastro no SEFIC foi realizado com sucesso!</p> <p>Seu C&oacute;digo &eacute; '+ str(UC)+'.</p> <p>&nbsp;</p> \
                    <p>Para submeter seu projeto &agrave; Chamada P&uacute;blica de Projetos 2021 acesse:</p> <p>&nbsp;</p> \
                    <p><a href="http://sefic.cpfl.com.br/Conta/LoginExterno">http://sefic.cpfl.com.br/Conta/LoginExterno</a></p> <p>&nbsp;</p> \
                    <p>Atenciosamente,</p> <p>&nbsp;</p> <p><strong>Comiss&atilde;o de Chamada P&uacute;blica de Projetos 2021</strong></p> \
                    <p><strong>CPFL Energia</strong></p></font></p>'
                    mail.SentOnBehalfOfName = 'peecpfl@cpfl.com.br'
                    mail.Send()
            
            #atualizar planilha com a informação de que já haviam registros
            else:
                book = openpyxl.load_workbook(path)
                sheet = book.active
                fontStyle = Font(name="Calibri", size=12, color=colors.BLACK)
                celula_confirmacao = sheet.cell(row=linha+3, column=23)
                celula_confirmacao.value = text
                celula_confirmacao.font = fontStyle
                book.save(path)
                book.close()
                
    linha=linha+1

ROOT = tk.Tk()

ROOT.withdraw()

messagebox.showinfo("Status do cadastro de UCs", "Ufa! Acabei aqui! Fiz o cadastro de "+ str(k)+" UCs")

driver.close()